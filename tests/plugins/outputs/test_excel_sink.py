"""Tests for Excel result sink."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from elspeth.plugins.outputs.excel import ExcelResultSink


@pytest.fixture
def mock_workbook():
    """Create a mock workbook."""
    mock_wb = MagicMock()
    mock_active_sheet = MagicMock()
    mock_wb.active = mock_active_sheet
    mock_wb.create_sheet.return_value = MagicMock()
    return mock_wb


@pytest.fixture
def sample_results():
    """Create sample results data."""
    return {
        "results": [
            {
                "row": {"id": 1, "name": "Alice"},
                "content": "Response 1",
                "score": 0.8,
            },
            {
                "row": {"id": 2, "name": "Bob"},
                "content": "Response 2",
                "score": 0.9,
            },
        ],
        "aggregates": {
            "mean_score": 0.85,
            "total_rows": 2,
        },
        "cost_summary": {
            "total": 0.01,
            "currency": "USD",
        },
    }


class TestExcelResultSinkInit:
    """Tests for ExcelResultSink initialization."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_init_with_defaults(self, mock_load):
        """Initializes with default values."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(base_path="/output")

        assert sink.base_path == Path("/output")
        assert sink.workbook_name is None
        assert sink.timestamped is True
        assert sink.results_sheet == "Results"
        assert sink.manifest_sheet == "Manifest"
        assert sink.aggregates_sheet == "Aggregates"
        assert sink.include_manifest is True
        assert sink.include_aggregates is True
        assert sink.on_error == "abort"

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_init_with_custom_options(self, mock_load):
        """Accepts custom options."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(
            base_path="/custom/output",
            workbook_name="my_workbook",
            timestamped=False,
            results_sheet="Data",
            manifest_sheet="Info",
            aggregates_sheet="Stats",
            include_manifest=False,
            include_aggregates=False,
        )

        assert sink.base_path == Path("/custom/output")
        assert sink.workbook_name == "my_workbook"
        assert sink.timestamped is False
        assert sink.results_sheet == "Data"
        assert sink.manifest_sheet == "Info"
        assert sink.aggregates_sheet == "Stats"
        assert sink.include_manifest is False
        assert sink.include_aggregates is False

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_init_rejects_invalid_on_error(self, mock_load):
        """Rejects invalid on_error value."""
        mock_load.return_value = MagicMock()
        with pytest.raises(ValueError, match="on_error must be 'abort'"):
            ExcelResultSink(base_path="/output", on_error="ignore")

    def test_init_requires_openpyxl(self):
        """Raises error when openpyxl is not available."""
        with patch(
            "elspeth.plugins.outputs.excel._load_workbook_dependencies",
            side_effect=RuntimeError("openpyxl not installed"),
        ), pytest.raises(RuntimeError, match="openpyxl"):
            ExcelResultSink(base_path="/output")


class TestExcelResultSinkWrite:
    """Tests for ExcelResultSink.write()."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_write_creates_workbook(self, mock_load, tmp_path, sample_results):
        """Write creates workbook file."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(base_path=tmp_path, timestamped=False)
        sink.write(sample_results, metadata={"experiment": "test_exp"})

        mock_workbook.save.assert_called_once()
        save_path = mock_workbook.save.call_args[0][0]
        assert "test_exp" in str(save_path)
        assert str(save_path).endswith(".xlsx")

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_write_with_timestamped_name(self, mock_load, tmp_path, sample_results):
        """Write adds timestamp to filename."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(base_path=tmp_path, timestamped=True)
        sink.write(sample_results, metadata={"experiment": "test"})

        save_path = mock_workbook.save.call_args[0][0]
        # Format: test_YYYYMMDDTHHMMSSz.xlsx
        assert "test_" in str(save_path)
        assert "T" in str(save_path)  # Timestamp includes T separator
        assert "Z.xlsx" in str(save_path)

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_write_custom_workbook_name(self, mock_load, tmp_path, sample_results):
        """Write uses custom workbook name."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="custom_name",
            timestamped=False,
        )
        sink.write(sample_results)

        save_path = mock_workbook.save.call_args[0][0]
        assert save_path == tmp_path / "custom_name.xlsx"

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_write_strips_xlsx_extension_from_name(self, mock_load, tmp_path, sample_results):
        """Write strips .xlsx from custom name before adding it."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="test.xlsx",
            timestamped=False,
        )
        sink.write(sample_results)

        save_path = mock_workbook.save.call_args[0][0]
        assert save_path == tmp_path / "test.xlsx"
        assert "test.xlsx.xlsx" not in str(save_path)


class TestExcelResultsSheet:
    """Tests for results sheet population."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_populates_results_with_headers(self, mock_load, tmp_path, sample_results):
        """Results sheet includes headers from all rows."""
        mock_sheet = MagicMock()
        appended_rows = []
        mock_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(base_path=tmp_path, timestamped=False)
        sink.write(sample_results)

        # First row should be headers
        headers = appended_rows[0]
        assert "content" in headers
        assert "score" in headers
        assert "row.id" in headers
        assert "row.name" in headers

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_populates_results_with_flattened_data(self, mock_load, tmp_path):
        """Results sheet flattens row data."""
        mock_sheet = MagicMock()
        appended_rows = []
        mock_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        results = {
            "results": [
                {"row": {"id": 1}, "content": "test", "score": 0.5},
            ],
        }

        sink = ExcelResultSink(base_path=tmp_path, timestamped=False)
        sink.write(results)

        # Second row should be data
        headers = appended_rows[0]
        data_row = appended_rows[1]
        data_dict = dict(zip(headers, data_row))

        assert data_dict["row.id"] == 1
        assert data_dict["content"] == "test"
        assert data_dict["score"] == 0.5

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_sets_results_sheet_title(self, mock_load, tmp_path):
        """Results sheet gets configured title."""
        mock_sheet = MagicMock()
        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            results_sheet="CustomResults",
            timestamped=False,
        )
        sink.write({"results": []})

        assert mock_sheet.title == "CustomResults"

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_empty_results_shows_placeholder(self, mock_load, tmp_path):
        """Empty results show 'no_results' placeholder."""
        mock_sheet = MagicMock()
        appended_rows = []
        mock_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = mock_sheet
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(base_path=tmp_path, timestamped=False)
        sink.write({"results": []})

        assert appended_rows[0] == ["no_results"]


class TestExcelManifestSheet:
    """Tests for manifest sheet population."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_creates_manifest_sheet(self, mock_load, tmp_path, sample_results):
        """Creates manifest sheet when enabled."""
        mock_manifest_sheet = MagicMock()
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = mock_manifest_sheet
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=True,
            timestamped=False,
        )
        sink.write(sample_results)

        mock_workbook.create_sheet.assert_any_call("Manifest")

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_skips_manifest_when_disabled(self, mock_load, tmp_path, sample_results):
        """Skips manifest sheet when disabled."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=False,
            include_aggregates=False,
            timestamped=False,
        )
        sink.write(sample_results)

        mock_workbook.create_sheet.assert_not_called()

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_manifest_includes_generated_at(self, mock_load, tmp_path, sample_results):
        """Manifest includes generation timestamp."""
        mock_manifest_sheet = MagicMock()
        appended_rows = []
        mock_manifest_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = mock_manifest_sheet
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=True,
            include_aggregates=False,
            timestamped=False,
        )
        sink.write(sample_results)

        keys = [row[0] for row in appended_rows]
        assert "generated_at" in keys

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_manifest_includes_row_count(self, mock_load, tmp_path, sample_results):
        """Manifest includes row count."""
        mock_manifest_sheet = MagicMock()
        appended_rows = []
        mock_manifest_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = mock_manifest_sheet
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=True,
            include_aggregates=False,
            timestamped=False,
        )
        sink.write(sample_results)

        # Find rows entry
        row_dict = {row[0]: row[1] for row in appended_rows if len(row) == 2}
        assert row_dict.get("rows") == 2

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_manifest_includes_cost_summary(self, mock_load, tmp_path, sample_results):
        """Manifest includes cost summary when present."""
        mock_manifest_sheet = MagicMock()
        appended_rows = []
        mock_manifest_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = mock_manifest_sheet
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=True,
            include_aggregates=False,
            timestamped=False,
        )
        sink.write(sample_results)

        keys = [row[0] for row in appended_rows]
        assert "cost_summary" in keys


class TestExcelAggregatesSheet:
    """Tests for aggregates sheet population."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_creates_aggregates_sheet(self, mock_load, tmp_path, sample_results):
        """Creates aggregates sheet when enabled and data present."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=False,
            include_aggregates=True,
            timestamped=False,
        )
        sink.write(sample_results)

        mock_workbook.create_sheet.assert_called_with("Aggregates")

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_skips_aggregates_when_disabled(self, mock_load, tmp_path, sample_results):
        """Skips aggregates sheet when disabled."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=False,
            include_aggregates=False,
            timestamped=False,
        )
        sink.write(sample_results)

        mock_workbook.create_sheet.assert_not_called()

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_skips_aggregates_when_no_data(self, mock_load, tmp_path):
        """Skips aggregates sheet when no aggregates data."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=False,
            include_aggregates=True,
            timestamped=False,
        )
        sink.write({"results": []})  # No aggregates

        mock_workbook.create_sheet.assert_not_called()

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_aggregates_includes_all_metrics(self, mock_load, tmp_path, sample_results):
        """Aggregates sheet includes all metrics."""
        mock_agg_sheet = MagicMock()
        appended_rows = []
        mock_agg_sheet.append.side_effect = lambda row: appended_rows.append(list(row))

        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = mock_agg_sheet
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            include_manifest=False,
            include_aggregates=True,
            timestamped=False,
        )
        sink.write(sample_results)

        keys = [row[0] for row in appended_rows if len(row) == 2]
        assert "mean_score" in keys
        assert "total_rows" in keys


class TestExcelFlattenResult:
    """Tests for result flattening."""

    def test_flattens_row_dict(self):
        """Flattens row dict with prefix."""
        entry = {
            "row": {"id": 1, "name": "test"},
            "content": "result",
        }
        flat = ExcelResultSink._flatten_result(entry)

        assert flat["row.id"] == 1
        assert flat["row.name"] == "test"
        assert flat["content"] == "result"

    def test_handles_nested_values_as_json(self):
        """Converts nested values to JSON strings."""
        entry = {
            "metrics": {"score": 0.5, "tokens": 100},
            "tags": ["a", "b"],
        }
        flat = ExcelResultSink._flatten_result(entry)

        assert flat["metrics"] == json.dumps({"score": 0.5, "tokens": 100}, sort_keys=True)
        assert flat["tags"] == json.dumps(["a", "b"], sort_keys=True)

    def test_handles_missing_row(self):
        """Handles entry without row key."""
        entry = {"content": "test", "score": 0.8}
        flat = ExcelResultSink._flatten_result(entry)

        assert flat["content"] == "test"
        assert flat["score"] == 0.8
        assert "row" not in flat


class TestExcelArtifacts:
    """Tests for artifact collection."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_collect_artifacts_after_write(self, mock_load, tmp_path, sample_results):
        """Collect artifacts returns workbook path."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="test",
            timestamped=False,
        )
        sink.write(sample_results, metadata={"security_level": "official"})
        artifacts = sink.collect_artifacts()

        assert "excel" in artifacts
        artifact = artifacts["excel"]
        assert artifact.path == str(tmp_path / "test.xlsx")
        assert artifact.type == "file/xlsx"
        assert artifact.security_level == "official"

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_collect_artifacts_clears_state(self, mock_load, tmp_path, sample_results):
        """Collect artifacts clears internal state."""
        mock_workbook = MagicMock()
        mock_workbook.active = MagicMock()
        mock_workbook.create_sheet.return_value = MagicMock()
        mock_load.return_value = lambda: mock_workbook

        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="test",
            timestamped=False,
        )
        sink.write(sample_results)

        # First collection
        artifacts = sink.collect_artifacts()
        assert "excel" in artifacts

        # Second collection should be empty
        artifacts2 = sink.collect_artifacts()
        assert artifacts2 == {}

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_collect_artifacts_empty_before_write(self, mock_load, tmp_path):
        """Collect artifacts returns empty dict before write."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(base_path=tmp_path)

        artifacts = sink.collect_artifacts()
        assert artifacts == {}


class TestExcelProducesConsumes:
    """Tests for artifact descriptors."""

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_produces_returns_xlsx_descriptor(self, mock_load, tmp_path):
        """Produces returns xlsx descriptor."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(base_path=tmp_path)

        descriptors = sink.produces()

        assert len(descriptors) == 1
        assert descriptors[0].name == "excel"
        assert descriptors[0].type == "file/xlsx"
        assert descriptors[0].persist is True

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_consumes_returns_empty(self, mock_load, tmp_path):
        """Consumes returns empty list."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(base_path=tmp_path)

        assert sink.consumes() == []

    @patch("elspeth.plugins.outputs.excel._load_workbook_dependencies")
    def test_finalize_returns_none(self, mock_load, tmp_path):
        """Finalize returns None."""
        mock_load.return_value = MagicMock()
        sink = ExcelResultSink(base_path=tmp_path)

        result = sink.finalize({})
        assert result is None


class TestExcelIntegration:
    """Integration tests with real openpyxl (if available)."""

    @pytest.fixture
    def skip_if_no_openpyxl(self):
        """Skip test if openpyxl is not installed."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_full_write_cycle(self, skip_if_no_openpyxl, tmp_path, sample_results):
        """Full write cycle creates valid Excel file."""
        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="integration_test",
            timestamped=False,
        )
        sink.write(sample_results, metadata={"experiment": "test"})

        # Verify file exists
        xlsx_path = tmp_path / "integration_test.xlsx"
        assert xlsx_path.exists()

        # Verify can be opened
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        sheet_names = wb.sheetnames

        assert "Results" in sheet_names
        assert "Manifest" in sheet_names
        assert "Aggregates" in sheet_names

    def test_writes_actual_data(self, skip_if_no_openpyxl, tmp_path):
        """Writes actual data to Excel cells."""
        results = {
            "results": [
                {"row": {"id": 1}, "content": "Hello", "score": 0.5},
            ],
            "aggregates": {"mean": 0.5},
        }
        sink = ExcelResultSink(
            base_path=tmp_path,
            workbook_name="data_test",
            timestamped=False,
        )
        sink.write(results)

        from openpyxl import load_workbook

        wb = load_workbook(tmp_path / "data_test.xlsx")
        results_sheet = wb["Results"]

        # Get all values
        rows = list(results_sheet.iter_rows(values_only=True))
        headers = rows[0]
        data = rows[1]

        # Find content column
        content_idx = headers.index("content")
        assert data[content_idx] == "Hello"
